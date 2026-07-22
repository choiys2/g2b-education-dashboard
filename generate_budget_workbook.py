#!/usr/bin/env python3
"""
컨소시엄 예산워크북 생성 엔진
사용법: python3 generate_budget_workbook.py config.json output.xlsx
config.json 구조는 config_example.json 참고.
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
SUBHEADER_FILL = PatternFill("solid", start_color="D9E1F2")
INPUT_FONT = Font(name=FONT_NAME, color="0000FF")
FORMULA_FONT = Font(name=FONT_NAME, color="000000")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_overview(wb, cfg):
    ws = wb.active
    ws.title = "01_개요"
    ws["A1"] = cfg["사업명"]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    rows = [
        ("권역", cfg["권역"]),
        ("주관기관", cfg["주관기관"]),
        ("사업기간", f'{cfg["사업기간"]["시작"]} ~ {cfg["사업기간"]["종료"]}'),
        ("총사업비(원)", cfg["총사업비"]),
        ("참여기관 수", len(cfg["참여기관"])),
    ]
    r = 3
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, bold=True)
        c = ws.cell(row=r, column=2, value=val)
        if label == "총사업비(원)":
            c.font = INPUT_FONT
            c.number_format = "#,##0"
        r += 1
    autosize(ws, [18, 40, 15, 15])
    return ws


def build_detail_budget(wb, cfg):
    ws = wb.create_sheet("03_세부예산")
    headers = ["카테고리", "세부항목", "단가(원)", "수량", "횟수", "금액(원)", "비고"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, 1, len(headers))

    r = 2
    cat_ranges = {}
    for cat in cfg["예산카테고리"]:
        start = r
        for item in cat["items"]:
            ws.cell(row=r, column=1, value=cat["category"])
            ws.cell(row=r, column=2, value=item["항목"])
            pc = ws.cell(row=r, column=3, value=item["단가"])
            pc.font = INPUT_FONT
            pc.number_format = "#,##0"
            qc = ws.cell(row=r, column=4, value=item["수량"])
            qc.font = INPUT_FONT
            hc = ws.cell(row=r, column=5, value=item["횟수"])
            hc.font = INPUT_FONT
            amt = ws.cell(row=r, column=6, value=f"=C{r}*D{r}*E{r}")
            amt.font = FORMULA_FONT
            amt.number_format = "#,##0"
            ws.cell(row=r, column=7, value=item.get("비고", ""))
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = BORDER
            r += 1
        cat_ranges[cat["category"]] = (start, r - 1)
    autosize(ws, [16, 24, 14, 8, 8, 16, 24])
    return ws, cat_ranges, r - 1


def build_summary(wb, cfg, cat_ranges, detail_last_row):
    ws = wb.create_sheet("02_예산총괄", 1)
    ws["A1"] = "예산 총괄"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    headers = ["카테고리", "금액(원)", "비중(%)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 3)

    r = 4
    cat_rows = []
    for cat, (s, e) in cat_ranges.items():
        ws.cell(row=r, column=1, value=cat)
        amt = ws.cell(row=r, column=2, value=f"='03_세부예산'!F{s}:F{e}")
        # SUM formula (avoid array literal issues -> use SUM explicitly)
        ws.cell(row=r, column=2, value=f"=SUM('03_세부예산'!F{s}:F{e})")
        ws.cell(row=r, column=2).number_format = "#,##0"
        ws.cell(row=r, column=2).font = FORMULA_FONT
        cat_rows.append(r)
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="합계").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})")
    ws.cell(row=total_row, column=2).number_format = "#,##0"
    ws.cell(row=total_row, column=2).font = Font(name=FONT_NAME, bold=True)

    for rr in cat_rows:
        ws.cell(row=rr, column=3, value=f"=B{rr}/$B${total_row}")
        ws.cell(row=rr, column=3).number_format = "0.0%"
    for c in range(1, 4):
        for rr in list(range(4, total_row + 1)):
            ws.cell(row=rr, column=c).border = BORDER

    ws.cell(row=total_row + 2, column=1, value="검증: 총사업비 대비 차액(원)").font = Font(name=FONT_NAME, italic=True)
    check = ws.cell(row=total_row + 2, column=2, value=f"='01_개요'!B6-B{total_row}")
    check.number_format = "#,##0"
    autosize(ws, [20, 18, 12])
    return ws, total_row


def build_org_share(wb, cfg, summary_total_row):
    ws = wb.create_sheet("04_참여기관_분담내역")
    ws["A1"] = "참여기관별 분담 내역"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    headers = ["참여기관", "분담률(%)", "분담금액(원)", "역할"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 4)
    r = 4
    for org in cfg["참여기관"]:
        ws.cell(row=r, column=1, value=org["기관명"])
        rc = ws.cell(row=r, column=2, value=org["분담률"])
        rc.number_format = "0.0%"
        rc.font = INPUT_FONT
        amt = ws.cell(row=r, column=3, value=f"=B{r}*'02_예산총괄'!$B${summary_total_row}")
        amt.number_format = "#,##0"
        amt.font = FORMULA_FONT
        ws.cell(row=r, column=4, value=org.get("역할", ""))
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    ws.cell(row=r, column=1, value="합계").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})").number_format = "0.0%"
    ws.cell(row=r, column=3, value=f"=SUM(C4:C{r-1})").number_format = "#,##0"
    autosize(ws, [22, 12, 18, 24])


def build_region_alloc(wb, cfg):
    if "지역배분" not in cfg or not cfg["지역배분"]:
        return
    ws = wb.create_sheet("05_지역별_배분")
    ws["A1"] = "지역별 배분"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    headers = ["지역코드", "지역명", "참여교사수", "1인당 배정예산(원)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 4)
    r = 4
    start = r
    for reg in cfg["지역배분"]:
        ws.cell(row=r, column=1, value=reg["지역코드"])
        ws.cell(row=r, column=2, value=reg["지역명"])
        tc = ws.cell(row=r, column=3, value=reg["참여교사수"])
        tc.font = INPUT_FONT
        r += 1
    end = r - 1
    ws.cell(row=r, column=2, value="합계").font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=r, column=3, value=f"=SUM(C{start}:C{end})")
    for rr in range(start, end + 1):
        ws.cell(row=rr, column=4, value=f"='02_예산총괄'!$B${{TOTAL_ROW}}/$C${r}").number_format = "#,##0"
    autosize(ws, [12, 16, 14, 20])


def generate(cfg_path, out_path):
    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)
    wb = Workbook()
    build_overview(wb, cfg)
    detail_ws, cat_ranges, detail_last = build_detail_budget(wb, cfg)
    summary_ws, total_row = build_summary(wb, cfg, cat_ranges, detail_last)
    build_org_share(wb, cfg, total_row)
    build_region_alloc(wb, cfg)
    # fix region alloc formula placeholder now that total_row known
    if "05_지역별_배분" in wb.sheetnames:
        ws = wb["05_지역별_배분"]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{TOTAL_ROW}" in cell.value:
                    cell.value = cell.value.replace("{TOTAL_ROW}", str(total_row))
    wb.save(out_path)
    print(f"saved: {out_path}")


if __name__ == "__main__":
    generate(sys.argv[1], sys.argv[2])
